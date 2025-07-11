import csv
import openpyxl
import os


def csv_to_excel(folder_path='data', excel_file='output.xlsx'):
    if not os.path.exists(folder_path):
        print(f"Папка '{folder_path}' не найдена.")
        return

    csv_files = [f for f in os.listdir(folder_path) if f.endswith('.csv')]

    if not csv_files:
        print(f"В папке '{folder_path}' нет CSV-файлов.")
        return


    wb = openpyxl.Workbook()


    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])


    for csv_file in csv_files:
        csv_path = os.path.join(folder_path, csv_file)

        sheet_name = os.path.splitext(csv_file)[0]

        ws = wb.create_sheet(title=sheet_name)

        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            for row in reader:
                ws.append(row)

    wb.save(excel_file)
    print(f"Excel-документ '{excel_file}' успешно создан из CSV-файлов в папке '{folder_path}'.")



if __name__ == "__main__":
    csv_to_excel(folder_path='data', excel_file='output.xlsx')